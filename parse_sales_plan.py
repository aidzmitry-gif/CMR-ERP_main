"""Парсер реальной таблицы «План по отделу продаж 2026 (план_факт).xlsx».

Извлекает по каждому месяцу: блоки менеджеров + строку «ОТДЕЛ ПРОДАЖ»,
с планом (столбец B) и фактом (столбец C) по каноническим показателям воронки.
Результат → rop-plan-real.json (для калькулятора/экрана «Планирование» РОП).

Запуск:  .venv\Scripts\python.exe parse_sales_plan.py
"""
from __future__ import annotations

import json

import openpyxl

SRC = r"D:/11 ERP Аккумуляторные/План по отделу продаж 2026 (план_факт).xlsx"
OUT = "rop-plan-real.json"

# Канонические показатели → ключевые слова (порядок важен: более узкие выше).
CANON = [
    ("deals_active",  ["дел на сегодня"]),
    ("calls",         ["холодных"]),
    ("leads",         ["лидов"]),
    ("invoices",      ["новых сделок", "счетов отправленных"]),
    ("carry_sum",     ["перешедшие на будущие периоды", "сумма"]),
    ("carry_count",   ["перешедшие на будущие периоды", "количество"]),
    ("carry_profit",  ["прибыль успешных сделок перешедших"]),
    ("conv",          ["конверсия"]),
    ("new_sum",       ["сумма новых сделок"]),
    ("tender_sum",    ["сумма тендера"]),
    ("tender_count",  ["тендер"]),
    ("wins_sum",      ["сумма успешных"]),
    ("wins_count",    ["успешные сделки"]),
    ("cash",          ["оплаты с ндс"]),
    ("shipments",     ["отгрузки"]),
    ("profit",        ["прибыль валовая"]),
    ("lost_sum",      ["сумма проигранных"]),
    ("lost_count",    ["проигранные сделки"]),
    ("avg_check",     ["средний чек"]),
    ("closed",        ["закрытые сделки"]),
    ("workdays",      ["количество рабочих дней"]),
]
STOP = {"менеджер", "план", "выручка", "вал прибыль", "что можно лучше?",
        "текущий месяц", "количество рабочих прошедщих в этом месяце"}


def num(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return round(float(v), 2)
    s = str(v).strip()
    if s == "" or s.startswith("#"):  # #REF! / #DIV/0!
        return None
    try:
        return round(float(s.replace(" ", "").replace(",", ".")), 2)
    except ValueError:
        return None


def match_indicator(label: str):
    low = label.lower()
    if low in STOP:
        return None
    for key, kws in CANON:
        if all(k in low for k in kws):
            return key
    return None


def is_block_header(ws, r) -> bool:
    """Заголовок блока (имя менеджера / ОТДЕЛ ПРОДАЖ): не показатель, и в ближайших
    6 строках есть индикатор «дел на сегодня» или «холодных звонков»."""
    a = ws.cell(r, 1).value
    if not a or not str(a).strip():
        return False
    low = str(a).strip().lower()
    if low in STOP or match_indicator(str(a).strip()):
        return False
    for rr in range(r + 1, min(r + 7, ws.max_row + 1)):
        nxt = ws.cell(rr, 1).value
        if nxt and ("дел на сегодня" in str(nxt).lower() or "холодных" in str(nxt).lower()):
            return True
    return False


def parse_block(ws, start, end):
    plan, fact = {}, {}
    for r in range(start, end):
        a = ws.cell(r, 1).value
        if not a or not str(a).strip():
            continue
        key = match_indicator(str(a).strip())
        if not key:
            continue
        p, f = num(ws.cell(r, 2).value), num(ws.cell(r, 3).value)
        if p is not None:
            plan[key] = p
        if f is not None:
            fact[key] = f
    return plan, fact


def parse_sheet(ws):
    headers = [r for r in range(1, ws.max_row + 1) if is_block_header(ws, r)]
    headers.append(ws.max_row + 1)
    blocks = []
    for i in range(len(headers) - 1):
        r = headers[i]
        name = str(ws.cell(r, 1).value).strip()
        plan, fact = parse_block(ws, r + 1, headers[i + 1])
        if plan or fact:
            blocks.append({"name": name, "plan": plan, "fact": fact})
    managers = [b for b in blocks if "отдел" not in b["name"].lower()]
    otdel = next((b for b in blocks if "отдел" in b["name"].lower()), None)
    return managers, otdel


def main():
    wb = openpyxl.load_workbook(SRC, data_only=True)
    months = []
    for ws in wb.worksheets:
        title = ws.title.strip()
        if title.lower().startswith("свод"):
            continue
        managers, otdel = parse_sheet(ws)
        if not managers and not otdel:
            continue
        months.append({"sheet": title, "managers": managers, "otdel": otdel})
    data = {"source": SRC, "currency": "BYN", "months": months}
    with open(OUT, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    # краткая сводка
    for m in months:
        names = ", ".join(x["name"] for x in m["managers"])
        prof = (m["otdel"] or {}).get("plan", {}).get("profit")
        print(f"{m['sheet']:>12} | менеджеров: {len(m['managers'])} ({names}) | отдел.прибыль_план={prof}")
    print(f"\n→ {OUT}: {len(months)} мес.")


if __name__ == "__main__":
    main()
